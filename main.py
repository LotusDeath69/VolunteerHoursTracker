import discord
from discord.ext import commands
from dotenv import load_dotenv
from discord.errors import DiscordException
import os 
from database import *
from datetime import timedelta
from openpyxl import Workbook


load_dotenv()


#Varibles
prefix = '?'
token = os.environ['TOKEN']
client = commands.Bot(command_prefix=prefix, help_command=None)


def addData(data):
  wb = Workbook()
  ws = wb.active
  ws['A1'] = 'Date'
  ws['B1'] = 'Hours'
  ws['C1'] = 'Total Hours'
  for i in data:
    row = list(i)
    del row[2]
    ws.append(row)
  wb.save("volunteerHours.xlsx")


def createEmbed(ign, title, value):
    embed = discord.Embed(
    title = title,
    colour = discord.Colour.blue()
  )
    embed.add_field(name=ign, value=value)
    embed.set_footer(text='Created by ThatBananaking')
    return embed 


@client.event
async def on_ready():
  await client.change_presence(status=discord.Status.online, activity=discord.Activity(
    type=discord.ActivityType.listening, name='change this to total hours'
  ))
  print('logged in')


@client.command()
async def add(ctx, *args):
  if discord.author.id != 466042357553430539:
    await ctx.reply('You cannot use this bot! L')
    return

  if len(args) > 2:
    await ctx.reply('Requires only `date, hours` input.')
  elif len(args) == 1:
    await ctx.reply('Requires the `hours` variable as well.')
  
  else:
    if ':' in args[1]:
      time = args[1].split(':')
      time = int(time[0]) * 60 + int(time[1])
    else:
      time = int(args[1]) * 60

    try:
      total_hours = int(logRetrive()[::-1][0][3]) + time
    except IndexError:
      total_hours = 0
    logAdd(args[0], str(time), str(total_hours))
    await ctx.reply(f'Congratulations! You now have {divmod(total_hours, 60)[0]}/100 hours.')


@client.command()
async def total(ctx):
  await ctx.reply(f"You have {divmod(int(logRetrive()[::-1][0][3]), 60)[0]}/100 hours.")


@client.command()
async def excel(ctx):
  addData(logRetrive())
  await ctx.send(file=discord.File(r'C:\Users\lotud\Documents\code\Tracker\volunteerHours.xlsx'))

client.run(token)

